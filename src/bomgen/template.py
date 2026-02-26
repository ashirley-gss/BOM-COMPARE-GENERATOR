"""Template handling for BOM Excel files."""

from pathlib import Path
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import BOM, BOMItem, BOMComparison


class BOMTemplate:
    """Handles Excel template operations for BOMs."""
    
    def __init__(self, template_path: Optional[Path] = None):
        """Initialize with optional template path."""
        if template_path is None:
            # Default to package template
            template_path = Path(__file__).parent.parent.parent / "templates" / "BOM_COMPARE_TEMPLATE.xlsx"
        self.template_path = template_path
    
    def create_workbook(self) -> openpyxl.Workbook:
        """Create a new workbook with default BOM structure."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOM"
        
        # Set up headers
        headers = ["Part Number", "Description", "Quantity", "Unit", "Reference Designator", "Notes"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set column widths
        column_widths = [20, 40, 12, 10, 20, 30]
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        return wb
    
    def write_bom(self, bom: BOM, output_path: Path) -> None:
        """Write a BOM to an Excel file."""
        wb = self.create_workbook()
        ws = wb.active
        
        # Write metadata
        ws.cell(row=1, column=1, value="BOM Name:")
        ws.cell(row=1, column=2, value=bom.name)
        ws.cell(row=2, column=1, value="Version:")
        ws.cell(row=2, column=2, value=bom.version)
        ws.cell(row=3, column=1, value="Date:")
        ws.cell(row=3, column=2, value=bom.date.strftime("%Y-%m-%d"))
        
        # Write headers
        headers = ["Part Number", "Description", "Quantity", "Unit", "Reference Designator", "Notes"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write items
        for row_idx, item in enumerate(bom.items, start=6):
            ws.cell(row=row_idx, column=1, value=item.part_number)
            ws.cell(row=row_idx, column=2, value=item.description)
            ws.cell(row=row_idx, column=3, value=item.quantity)
            ws.cell(row=row_idx, column=4, value=item.unit)
            ws.cell(row=row_idx, column=5, value=item.reference_designator or "")
            ws.cell(row=row_idx, column=6, value=item.notes or "")
        
        wb.save(output_path)
    
    def write_comparison(self, comparison: BOMComparison, output_path: Path) -> None:
        """Write a BOM comparison to an Excel file."""
        wb = openpyxl.Workbook()
        
        # Summary sheet
        summary_ws = wb.active
        summary_ws.title = "Summary"
        summary_ws.cell(row=1, column=1, value="BOM Comparison Summary")
        summary_ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        summary_ws.cell(row=3, column=1, value="BOM 1:")
        summary_ws.cell(row=3, column=2, value=comparison.bom1.name)
        summary_ws.cell(row=4, column=1, value="BOM 2:")
        summary_ws.cell(row=4, column=2, value=comparison.bom2.name)
        
        summary_ws.cell(row=6, column=1, value="Added Items:")
        summary_ws.cell(row=6, column=2, value=len(comparison.added_items))
        summary_ws.cell(row=7, column=1, value="Removed Items:")
        summary_ws.cell(row=7, column=2, value=len(comparison.removed_items))
        summary_ws.cell(row=8, column=1, value="Modified Items:")
        summary_ws.cell(row=8, column=2, value=len(comparison.modified_items))
        summary_ws.cell(row=9, column=1, value="Unchanged Items:")
        summary_ws.cell(row=9, column=2, value=len(comparison.unchanged_items))
        
        # Added items sheet
        if comparison.added_items:
            added_ws = wb.create_sheet("Added Items")
            self._write_items_sheet(added_ws, comparison.added_items, "Added Items")
        
        # Removed items sheet
        if comparison.removed_items:
            removed_ws = wb.create_sheet("Removed Items")
            self._write_items_sheet(removed_ws, comparison.removed_items, "Removed Items")
        
        # Modified items sheet
        if comparison.modified_items:
            modified_ws = wb.create_sheet("Modified Items")
            headers = ["Part Number", "Field", "Old Value", "New Value"]
            for col_idx, header in enumerate(headers, start=1):
                cell = modified_ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            row = 2
            for old_item, new_item in comparison.modified_items:
                if old_item.quantity != new_item.quantity:
                    modified_ws.cell(row=row, column=1, value=old_item.part_number)
                    modified_ws.cell(row=row, column=2, value="Quantity")
                    modified_ws.cell(row=row, column=3, value=old_item.quantity)
                    modified_ws.cell(row=row, column=4, value=new_item.quantity)
                    row += 1
                if old_item.description != new_item.description:
                    modified_ws.cell(row=row, column=1, value=old_item.part_number)
                    modified_ws.cell(row=row, column=2, value="Description")
                    modified_ws.cell(row=row, column=3, value=old_item.description)
                    modified_ws.cell(row=row, column=4, value=new_item.description)
                    row += 1
                if old_item.unit != new_item.unit:
                    modified_ws.cell(row=row, column=1, value=old_item.part_number)
                    modified_ws.cell(row=row, column=2, value="Unit")
                    modified_ws.cell(row=row, column=3, value=old_item.unit)
                    modified_ws.cell(row=row, column=4, value=new_item.unit)
                    row += 1
        
        wb.save(output_path)
    
    def _write_items_sheet(self, ws: openpyxl.worksheet.worksheet.Worksheet, items: list[BOMItem], title: str) -> None:
        """Helper method to write items to a worksheet."""
        headers = ["Part Number", "Description", "Quantity", "Unit", "Reference Designator", "Notes"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for row_idx, item in enumerate(items, start=2):
            ws.cell(row=row_idx, column=1, value=item.part_number)
            ws.cell(row=row_idx, column=2, value=item.description)
            ws.cell(row=row_idx, column=3, value=item.quantity)
            ws.cell(row=row_idx, column=4, value=item.unit)
            ws.cell(row=row_idx, column=5, value=item.reference_designator or "")
            ws.cell(row=row_idx, column=6, value=item.notes or "")
