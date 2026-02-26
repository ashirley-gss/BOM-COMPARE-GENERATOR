"""Command-line interface for BOM generator."""

from pathlib import Path
from typing import Optional, List, Dict, Any
import typer
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import BOM, BOMItem, BOMComparison
from .template import BOMTemplate

app = typer.Typer(help="BOM Generator - A tool for generating and comparing Bill of Materials")

TEMPLATE_HEADERS = [
    "PartNo", "Revision", "Description", "AltDescription1", "AltDescription2", "DescExtra",
    "Quantity", "IssueUM", "ConsumptionConv", "UM", "Cost", "Source", "Drawing", "Leadtime",
    "Level", "Location", "Memo1", "Memo2", "Parent", "Productline", "Sequence", "SortCode",
    "Tag", "Category", "BomComplete", "BomComments", "Router"
]

REQUIRED_FIELDS = {"PartNo", "Quantity", "Parent", "Sequence"}


def load_template_headers(ws) -> List[str]:
    """Load headers from the first row of the worksheet."""
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    return headers


def append_row_by_headers(ws, headers: List[str], row_data: Dict[str, Any]) -> None:
    """Append a row to the worksheet based on header order, validating required fields."""
    # Validate required fields (allow empty string for Parent to avoid "descendant of itself" in BOM Compare)
    def is_missing(f: str) -> bool:
        val = row_data.get(f)
        if f == "Parent" and val == "":
            return False
        return val in (None, "", [])
    missing = [f for f in REQUIRED_FIELDS if is_missing(f)]
    if missing:
        raise ValueError(f"Missing required fields: {missing}")

    # Build row in correct column order
    values = [row_data.get(h, None) for h in headers]
    ws.append(values)


@app.command()
def generate(
    parent_part: str = typer.Option(..., prompt=True, help="Parent BOM part number"),
    child_count: int = typer.Option(2, prompt=True, help="How many child components?"),
    template_path: Optional[Path] = typer.Option(None, "--template", "-t", help="Template Excel file path"),
    output_path: Path = typer.Option(Path("output_bom.xlsx"), "--output", "-o", help="Output Excel file path")
):
    """Generate a BOM from a template by prompting for component details."""
    # Default template path relative to package
    if template_path is None:
        template_path = Path(__file__).parent.parent.parent / "templates" / "BOM_COMPARE_TEMPLATE.xlsx"
    
    if not template_path.exists():
        raise FileNotFoundError(f"Template file not found: {template_path}")
    
    wb = openpyxl.load_workbook(template_path)
    ws = wb["Template"]

    headers = load_template_headers(ws)

    # sanity check: template matches expected columns
    if headers != TEMPLATE_HEADERS:
        raise RuntimeError(
            "Template headers do not match expected format.\n"
            f"Found: {headers}\nExpected: {TEMPLATE_HEADERS}"
        )

    for i in range(child_count):
        typer.echo(f"\n--- Child #{i+1} ---")
        partno = typer.prompt("Child PartNo")
        desc = typer.prompt("Description", default="")
        qty = float(typer.prompt("Quantity", default="1"))
        source = typer.prompt("Source (1-char code)", default="")  # you define codes
        um = typer.prompt("UM", default="EA")

        row = {
            "Parent": parent_part,
            "Sequence": i + 1,
            "PartNo": partno,
            "Description": desc,
            "Quantity": qty,
            "Source": source,
            "UM": um,
        }

        append_row_by_headers(ws, headers, row)

    wb.save(output_path)
    typer.echo(f"\nSaved: {output_path.resolve()}")


@app.command()
def compare(
    bom1_file: Path = typer.Argument(..., help="First BOM file path"),
    bom2_file: Path = typer.Argument(..., help="Second BOM file path"),
    output_file: Path = typer.Option(..., "--output", "-o", help="Output comparison Excel file path"),
):
    """Compare two BOM files and generate a comparison report."""
    typer.echo(f"Comparing {bom1_file} and {bom2_file}")
    # TODO: Implement comparison logic
    typer.echo(f"Comparison report generated: {output_file}")


@app.command()
def create_template(
    output_file: Path = typer.Option(..., "--output", "-o", help="Output template file path"),
):
    """Create a blank BOM template Excel file with the correct headers."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    
    # Write headers
    for col_idx, header in enumerate(TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Set column widths
    for col_idx in range(1, len(TEMPLATE_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    
    wb.save(output_file)
    typer.echo(f"Template created: {output_file}")


if __name__ == "__main__":
    app()
