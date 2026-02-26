"""Streamlit web UI for BOM generator."""

import streamlit as st
from pathlib import Path
from typing import List, Dict, Any, Set
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io
import sys

# Ensure bomgen package is importable (add src dir to path when running as script)
current_file = Path(__file__).resolve()
project_root = current_file.parent.parent.parent
src_dir = project_root / "src"
if src_dir.exists() and str(src_dir) not in sys.path:
    sys.path.insert(0, str(src_dir))
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

# Import from cli module (bomgen lives under src/)
TEMPLATE_HEADERS = None
REQUIRED_FIELDS = None
load_template_headers = None
append_row_by_headers = None

try:
    from bomgen.cli import TEMPLATE_HEADERS, REQUIRED_FIELDS, load_template_headers, append_row_by_headers
except ImportError:
    try:
        from .cli import TEMPLATE_HEADERS, REQUIRED_FIELDS, load_template_headers, append_row_by_headers
    except ImportError:
        # Fallback: define inline so we never exec cli.py (which uses relative imports)
        TEMPLATE_HEADERS = [
            "PartNo", "Revision", "Description", "AltDescription1", "AltDescription2", "DescExtra",
            "Quantity", "IssueUM", "ConsumptionConv", "UM", "Cost", "Source", "Drawing", "Leadtime",
            "Level", "Location", "Memo1", "Memo2", "Parent", "Productline", "Sequence", "SortCode",
            "Tag", "Category", "BomComplete", "BomComments", "Router"
        ]
        REQUIRED_FIELDS = {"PartNo", "Quantity", "Parent", "Sequence"}

        def load_template_headers(ws):
            return [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

        def append_row_by_headers(ws, headers, row_data):
            def is_missing(f):
                val = row_data.get(f)
                if f == "Parent" and val == "":
                    return False
                return val in (None, "", [])
            missing = [f for f in REQUIRED_FIELDS if is_missing(f)]
            if missing:
                raise ValueError(f"Missing required fields: {missing}")
            values = [row_data.get(h, None) for h in headers]
            ws.append(values)

# Random data generator
try:
    from bomgen.random_data import random_child_rows, reset_part_number_counter
except ImportError:
    try:
        from .random_data import random_child_rows, reset_part_number_counter
    except ImportError:
        random_child_rows = None
        reset_part_number_counter = None

# Category and Source dropdowns: display label -> value in file (per user spec)
CATEGORY_OPTIONS = [
    ("Normal", ""),
    ("Phantom", "P"),
    ("Exclude", "X"),
    ("Reference", "R"),
    ("Setup", "1"),
]
SOURCE_OPTIONS = [
    ("—", ""),  # no default for component/sub-component manual input
    ("Purchase to Stock", "P"),
    ("Purchase to Job", "J"),
    ("Manufactured to Stock", "M"),
    ("Manufactured to Job", "F"),
    ("Consign to Stock", "C"),
    ("Consign to Job", "G"),
]
CATEGORY_DISPLAY_TO_VALUE = dict(CATEGORY_OPTIONS)
SOURCE_DISPLAY_TO_VALUE = dict(SOURCE_OPTIONS)
SOURCE_VALUE_TO_LABEL = {v: k for k, v in SOURCE_OPTIONS}  # for messages
CATEGORY_LABELS = [t[0] for t in CATEGORY_OPTIONS]
SOURCE_LABELS = [t[0] for t in SOURCE_OPTIONS]


def _apply_random_row_defaults(rows, manufactured_count, apply_revision_to_all, apply_location_to_all, parent_revision, parent_location):
    """Set UM=EA, Location/Revision from parent if apply, Category=Normal, Source (first N=F rest J), Productline CP/CM."""
    for i, row in enumerate(rows):
        row["UM"] = "EA"
        row["Location"] = (parent_location or None) if apply_location_to_all else None
        row["Revision"] = (parent_revision or None) if apply_revision_to_all else None
        row["Category"] = ""  # Normal
        row["Source"] = "F" if i < manufactured_count else "J"
        row["Productline"] = "CP" if row["Source"] == "F" else "CM"


def _validate_category_source(category_value, source_value, part_label=""):
    """Return error message if Category/Source combination is invalid. Phantom requires M or F; Exclude requires P."""
    if not category_value:
        return None
    if category_value == "P":  # Phantom
        if source_value not in ("M", "F"):
            return f"{part_label}: Phantom must have Source of Manufactured to Stock (M) or Manufactured to Job (F)."
    if category_value == "X":  # Exclude
        if source_value != "P":
            return f"{part_label}: Exclude must have Source of Purchase to Stock (P)."
    return None


def main():
    # Reset sequential part number counter (A001, A002, ...) at start of each run
    if reset_part_number_counter is not None:
        reset_part_number_counter()
    st.set_page_config(page_title="BOM Compare File Generator", page_icon="😀", layout="wide")
    st.title("😀 BOM Compare File Generator")
    st.markdown("Generate a Bill of Material .xlsx file with user defined or randomly generated parts to be used in BOM Compare.")

    # Sidebar: template selection
    with st.sidebar:
        st.header("Template Settings")
        template_option = st.radio(
            "Template Source",
            ["Use Default Template", "Upload Template", "Use Custom Path"],
            help="Choose how to load the template file",
        )
        template_path = None
        if template_option == "Use Default Template":
            default_path = Path(__file__).parent.parent.parent / "templates" / "BOM_COMPARE_TEMPLATE.xlsx"
            if default_path.exists():
                template_path = default_path
                st.success("Using default template")
            else:
                st.error("Default template not found")
                st.info("Upload a template or use a custom path")
        elif template_option == "Upload Template":
            uploaded_file = st.file_uploader("Upload Template Excel File", type=["xlsx", "xls"])
            if uploaded_file:
                temp_path = Path("temp_template.xlsx")
                with open(temp_path, "wb") as f:
                    f.write(uploaded_file.getbuffer())
                template_path = temp_path
                st.success("Template uploaded")
        elif template_option == "Use Custom Path":
            custom_path = st.text_input("Template File Path")
            if custom_path:
                template_path = Path(custom_path)
                if template_path.exists():
                    st.success("Template path valid")
                else:
                    st.error("Template file not found")

    if not template_path or not template_path.exists():
        st.info("👈 Select or upload a template file in the sidebar to begin.")
        return

    try:
        wb = openpyxl.load_workbook(template_path)
        if "Template" not in wb.sheetnames:
            st.error("Template file must contain a sheet named 'Template'")
            return
        ws = wb["Template"]
        headers = load_template_headers(ws)
        if headers != TEMPLATE_HEADERS:
            st.warning("Template headers don't match expected format.")
            with st.expander("Show Header Mismatch"):
                st.write("Found:", headers)
                st.write("Expected:", TEMPLATE_HEADERS)
            if not st.button("Continue Anyway"):
                return
    except Exception as e:
        st.error(f"Error loading template: {e}")
        return

    # --- Parent Part (first row, Level = 0) ---
    st.header("1. Top Parent Part (Level 0)")
    st.caption("This part will be written as the first data row in the .xlsx with Level = 0.")
    st.subheader("BOM structure")
    st.caption("Expand or collapse each level. Enter parts in a hierarchy like BOM Compare (Parent → Level 1 → Level 2 → Level 3).")
    with st.expander("📄 Parent part", expanded=True):
        c1, c2, c3, c4, c5 = st.columns(5)
        with c1:
            parent_partno = st.text_input("Parent PartNo *", key="parent_partno", placeholder="")
            parent_description = st.text_input("Description *", key="parent_desc", placeholder="")
            parent_quantity = st.number_input("Quantity *", min_value=0.0, value=1.0, key="parent_qty", step=1.0, format="%.2f")
            parent_um = st.text_input("UM", key="parent_um", value="EA")
            parent_productline = st.text_input("Productline", key="parent_productline", value="FG", placeholder="")
        with c2:
            parent_revision = st.text_input("Revision", key="parent_revision", placeholder="")
            parent_altdesc1 = st.text_input("AltDescription1", key="parent_altdesc1", placeholder="")
            parent_altdesc2 = st.text_input("AltDescription2", key="parent_altdesc2", placeholder="")
            parent_descextra = st.text_input("DescExtra", key="parent_descextra", placeholder="")
            parent_issue_um = st.text_input("IssueUM", key="parent_issue_um", placeholder="")
        with c3:
            parent_consumption = st.text_input("ConsumptionConv", key="parent_consumption", placeholder="")
            parent_cost = st.number_input("Cost", min_value=0.0, value=None, step=0.01, format="%.2f", key="parent_cost")
            parent_source_label = st.selectbox("Source", options=SOURCE_LABELS, index=SOURCE_LABELS.index("Manufactured to Stock"), key="parent_source")
            parent_drawing = st.text_input("Drawing", key="parent_drawing", placeholder="")
            parent_leadtime = st.text_input("Leadtime", key="parent_leadtime", placeholder="")
        with c4:
            parent_location = st.text_input("Location", key="parent_location", placeholder="")
            parent_memo1 = st.text_input("Memo1", key="parent_memo1", placeholder="")
            parent_memo2 = st.text_input("Memo2", key="parent_memo2", placeholder="")
            parent_sortcode = st.text_input("SortCode", key="parent_sortcode", placeholder="")
            parent_tag = st.text_input("Tag", key="parent_tag", placeholder="")
        with c5:
            parent_category_label = st.selectbox("Category", options=CATEGORY_LABELS, key="parent_category")
            parent_bomcomplete = st.text_input("BomComplete", key="parent_bomcomplete", placeholder="")
            parent_bomcomments = st.text_input("BomComments", key="parent_bomcomments", placeholder="")
            parent_router = st.text_input("Router", key="parent_router", placeholder="")
        apply_revision_to_all = st.checkbox(
            "Apply Revision to all Component and Sub-Component Parts",
            value=False,
            help="When checked, the Revision value above is applied to every Level 1, 2, and 3 part (manual and random).",
            key="apply_rev_all"
        )
        apply_location_to_all = st.checkbox(
            "Apply Location to all Component and Sub-Component Parts",
            value=False,
            help="When checked, the Location value above is applied to every Level 1, 2, and 3 part (manual and random).",
            key="apply_loc_all"
        )
        sequence_increment_options = [1, 10, 100, 1000, 10000]
        sequence_increment = st.selectbox(
            "Sequence increment",
            options=sequence_increment_options,
            index=2,
            format_func=lambda x: str(x),
            help="Sequence values start at this number for the first part of each level and increase by this amount (e.g. 100, 200, 300). Parent stays 0; each level resets.",
            key="sequence_increment"
        )

    parent_row = None
    if parent_partno and parent_description is not None:
        parent_row = {h: None for h in TEMPLATE_HEADERS}
        parent_row["PartNo"] = parent_partno
        parent_row["Description"] = parent_description
        parent_row["Quantity"] = parent_quantity
        parent_row["UM"] = parent_um or "EA"
        parent_row["Productline"] = parent_productline or "FG"
        parent_row["Level"] = 0
        # Leave Parent blank for top-level part so BOM Compare does not treat it as "descendant of itself"
        parent_row["Parent"] = None
        parent_row["Sequence"] = 0
        # Remaining fields from form (blank if user left empty)
        parent_row["Revision"] = parent_revision or None
        parent_row["AltDescription1"] = parent_altdesc1 or None
        parent_row["AltDescription2"] = parent_altdesc2 or None
        parent_row["DescExtra"] = parent_descextra or None
        parent_row["IssueUM"] = parent_issue_um or None
        parent_row["ConsumptionConv"] = parent_consumption or None
        parent_row["Cost"] = round(parent_cost, 2) if parent_cost is not None else None
        parent_row["Source"] = SOURCE_DISPLAY_TO_VALUE.get(parent_source_label) or None
        parent_row["Drawing"] = parent_drawing or None
        parent_row["Leadtime"] = parent_leadtime or None
        parent_row["Location"] = parent_location or None
        parent_row["Memo1"] = parent_memo1 or None
        parent_row["Memo2"] = parent_memo2 or None
        parent_row["SortCode"] = parent_sortcode or None
        parent_row["Tag"] = parent_tag or None
        _cat_val = CATEGORY_DISPLAY_TO_VALUE.get(parent_category_label, "")
        parent_row["Category"] = _cat_val or None
        parent_row["BomComplete"] = parent_bomcomplete or None
        parent_row["BomComments"] = parent_bomcomments or None
        parent_row["Router"] = parent_router or None

    if not parent_partno:
        st.info("Enter the parent part number to continue.")
        return

    # --- Level 1: tree-style expander ---
    with st.expander("📁 Level 1 — Components", expanded=True):
        st.markdown('<div style="margin-left: 1rem; border-left: 2px solid #888; padding-left: 0.75rem;">', unsafe_allow_html=True)
        use_random_children = st.checkbox("Randomly generate child components", value=False, help="Check to auto-generate part numbers and data instead of entering manually.", key="use_random_children")
        child_count = st.number_input("Number of child components", min_value=1, max_value=100, value=2, step=1, key="child_count")
        child_data = []
        level1_random_fields_set = None  # set below when random L1 (for preview)
        if use_random_children:
            st.subheader("Fields to populate (random generation)")
            st.caption("Select which columns to fill with random data. UM, Location, Revision, Category, Source, and Product Line will default based on user selections above.")
            l1_manufactured_count = st.number_input(
                "How many components to have Source of Manufactured to Job?",
                min_value=0,
                max_value=int(child_count),
                value=0,
                step=1,
                key="l1_manufactured_count",
            )
            multiselect_options = [f for f in TEMPLATE_HEADERS if f not in ("Parent", "Sequence", "Level")]
            fields_for_random = st.multiselect(
                "Fields to populate",
                options=multiselect_options,
                default=["PartNo", "Description", "Quantity", "Cost"],
                key="fields_random",
            )
            fields_set = set(fields_for_random) | {"Parent", "Sequence", "Level", "PartNo", "Quantity"}
            level1_random_fields_set = fields_set
            if random_child_rows:
                child_data = random_child_rows(parent_partno, child_count, fields_to_populate=fields_set, level=1)
                _apply_random_row_defaults(child_data, int(l1_manufactured_count), apply_revision_to_all, apply_location_to_all, parent_revision, parent_location)
            else:
                st.error("Random data module not available.")
                child_data = []
        else:
            level1_random_fields_set = None
            st.subheader("Enter child component details")
            for i in range(child_count):
                with st.expander(f"Child #{i+1}", expanded=(i < 2)):
                    cols = st.columns([2, 1, 2, 1, 1, 1])  # narrower Qty/UM, wider Source
                    with cols[0]:
                        c_partno = st.text_input("PartNo *", key=f"c_partno_{i}")
                        c_desc = st.text_input("Description", key=f"c_desc_{i}")
                    with cols[1]:
                        c_qty = st.number_input("Quantity *", min_value=0.0, value=1.0, key=f"c_qty_{i}", step=1.0, format="%.2f")
                        c_um = st.text_input("UM", value="EA", key=f"c_um_{i}")
                    with cols[2]:
                        c_source_label = st.selectbox("Source", options=SOURCE_LABELS, key=f"c_source_{i}")
                        c_pl = st.text_input("Productline", key=f"c_pl_{i}", placeholder="")
                    with cols[3]:
                        rev_default = (parent_revision or "") if apply_revision_to_all else ""
                        c_revision = st.text_input("Revision", key=f"c_rev_{i}_ar{apply_revision_to_all}_p{parent_revision or ''}", value=rev_default)
                        c_cost = st.number_input("Cost", min_value=0.0, value=None, step=0.01, format="%.2f", key=f"c_cost_{i}")
                    with cols[4]:
                        loc_default = (parent_location or "") if apply_location_to_all else ""
                        c_location = st.text_input("Location", key=f"c_loc_{i}_al{apply_location_to_all}_p{parent_location or ''}", value=loc_default)
                        c_category_label = st.selectbox("Category", options=CATEGORY_LABELS, key=f"c_cat_{i}")
                    with cols[5]:
                        c_leadtime = st.text_input("Leadtime", key=f"c_lt_{i}")
                        c_drawing = st.text_input("Drawing", key=f"c_draw_{i}")
                    with st.expander("Additional fields", expanded=False):
                        ec1, ec2, ec3 = st.columns(3)
                        with ec1:
                            c_altdesc1 = st.text_input("AltDescription1", key=f"c_altdesc1_{i}", placeholder="")
                            c_altdesc2 = st.text_input("AltDescription2", key=f"c_altdesc2_{i}", placeholder="")
                            c_descextra = st.text_input("DescExtra", key=f"c_descextra_{i}", placeholder="")
                            c_issue_um = st.text_input("IssueUM", key=f"c_issue_um_{i}", placeholder="")
                            c_consumption = st.text_input("ConsumptionConv", key=f"c_consumption_{i}", placeholder="")
                        with ec2:
                            c_memo1 = st.text_input("Memo1", key=f"c_memo1_{i}", placeholder="")
                            c_memo2 = st.text_input("Memo2", key=f"c_memo2_{i}", placeholder="")
                            c_sortcode = st.text_input("SortCode", key=f"c_sortcode_{i}", placeholder="")
                            c_tag = st.text_input("Tag", key=f"c_tag_{i}", placeholder="")
                        with ec3:
                            c_bomcomplete = st.text_input("BomComplete", key=f"c_bomcomplete_{i}", placeholder="")
                            c_bomcomments = st.text_input("BomComments", key=f"c_bomcomments_{i}", placeholder="")
                            c_router = st.text_input("Router", key=f"c_router_{i}", placeholder="")
                    child_data.append({
                        "PartNo": c_partno,
                        "Revision": c_revision or None,
                        "Description": c_desc or None,
                        "AltDescription1": c_altdesc1 or None,
                        "AltDescription2": c_altdesc2 or None,
                        "DescExtra": c_descextra or None,
                        "Quantity": c_qty,
                        "IssueUM": c_issue_um or None,
                        "ConsumptionConv": c_consumption or None,
                        "UM": c_um or "EA",
                        "Cost": round(c_cost, 2) if c_cost is not None else None,
                        "Source": SOURCE_DISPLAY_TO_VALUE.get(c_source_label) or None,
                        "Drawing": c_drawing or None,
                        "Leadtime": c_leadtime or None,
                        "Level": 1,
                        "Location": c_location or None,
                        "Memo1": c_memo1 or None,
                        "Memo2": c_memo2 or None,
                        "Parent": parent_partno,
                        "Productline": c_pl or None,
                        "Sequence": (i + 1) * sequence_increment,
                        "SortCode": c_sortcode or None,
                        "Tag": c_tag or None,
                        "Category": CATEGORY_DISPLAY_TO_VALUE.get(c_category_label) or None,
                        "BomComplete": c_bomcomplete or None,
                        "BomComments": c_bomcomments or None,
                        "Router": c_router or None,
                    })
        st.markdown("</div>", unsafe_allow_html=True)

    level1_partnos = [c["PartNo"] for c in child_data if c.get("PartNo")]

    # Level 2 parent options: only Level 1 parts that are Manufactured (M or F)
    level1_partnos_manufactured = [pno for pno in level1_partnos if next((c.get("Source") for c in child_data if c.get("PartNo") == pno), None) in ("M", "F")]

    # --- Level 2 sub-components: tree-style expander ---
    level2_config = []
    level2_manual_partnos = []
    random_l2_per_l1 = None  # {count, fields} or None
    random_l3_per_l2 = None  # {count, fields} or None
    # Block Level 2 if no Level 1 part has a Manufactured Source (M or F)
    any_l1_manufactured = any(c.get("Source") in ("M", "F") for c in child_data if c.get("PartNo"))
    block_l2 = level1_partnos and not any_l1_manufactured

    if level1_partnos:
        with st.expander("📁 Level 2 — Sub-components", expanded=True):
            st.markdown('<div style="margin-left: 2rem; border-left: 2px solid #888; padding-left: 0.75rem;">', unsafe_allow_html=True)
            st.caption("Add parts that belong under a Level 1 part. Use the option below and/or the groups.")
            if block_l2:
                st.warning(
                    "No Level 1 part has a **Manufactured** Source. At least one part must be **Manufactured to Stock** or **Manufactured to Job** to use as a parent for Level 2 Sub-Components. "
                    "Change a part's Source above to a Manufactured option."
                )
                num_l2_groups = st.number_input("Number of Level 2 groups", min_value=0, max_value=0, value=0, step=1, key="num_l2_groups", disabled=True)
            else:
                # Randomly generate sub-components for each Level 1 part
                with st.expander("Randomly generate Level 2 Sub-Components for each Manufactured Level 1 Part", expanded=False):
                    use_random_l2_per_l1 = st.checkbox("For each Manufactured Level 1 Part, randomly generate Level 2 Sub-Components", key="random_l2_per_l1")
                    if use_random_l2_per_l1:
                        count_random_l2 = st.number_input("Number of Level 2 parts per Level 1 part", min_value=1, max_value=30, value=2, key="count_random_l2")
                        l2_per_l1_manufactured = st.number_input("How many sub-components to have Source of Manufactured to Job?", min_value=0, max_value=int(count_random_l2), value=0, step=1, key="l2_per_l1_manufactured")
                        multiselect_options_l2 = [f for f in TEMPLATE_HEADERS if f not in ("Parent", "Sequence", "Level")]
                        fields_random_l2 = st.multiselect("Fields to populate", options=multiselect_options_l2, default=["PartNo", "Description", "Quantity", "Cost"], key="fields_random_l2")
                        random_l2_per_l1 = {"count": count_random_l2, "fields": fields_random_l2, "manufactured_count": min(int(l2_per_l1_manufactured), count_random_l2)}
                num_l2_groups = st.number_input("Number of Level 2 groups", min_value=0, max_value=20, value=0, step=1, key="num_l2_groups")
            random_all_l2 = False
            if not block_l2 and num_l2_groups > 0:
                random_all_l2 = st.checkbox("Randomly generate all Level 2 groups. Note: Each group will have 2 Parts.", key="random_all_l2")
            multiselect_options_l2 = [f for f in TEMPLATE_HEADERS if f not in ("Parent", "Sequence", "Level")]
            if random_all_l2:
                l2_all_manufactured = st.number_input("How many sub-components per group to have Source of Manufactured to Job?", min_value=0, max_value=2, value=0, step=1, key="l2_all_manufactured")
                fields_all_l2 = st.multiselect("Fields to populate (all Level 2 groups)", options=multiselect_options_l2, default=["PartNo", "Description", "Quantity", "Cost"], key="fields_all_l2")
                for g in range(num_l2_groups):
                    parent_l1 = level1_partnos_manufactured[g % len(level1_partnos_manufactured)]
                    level2_config.append({"parent": parent_l1, "count": 2, "random": True, "fields": fields_all_l2, "manufactured_count": min(int(l2_all_manufactured), 2)})
            else:
                for g in range(num_l2_groups):
                    with st.expander(f"Level 2 group {g+1}"):
                        parent_l1 = st.selectbox("Parent (Level 1 part)", level1_partnos_manufactured, index=0, key=f"l2_parent_{g}")
                        count_l2 = st.number_input("Number of parts", min_value=1, max_value=50, value=2, key=f"l2_count_{g}")
                        use_random_l2 = st.checkbox("Randomly generate this group", key=f"l2_random_{g}")
                        if use_random_l2:
                            l2_grp_manufactured = st.number_input("How many sub-components to have Source of Manufactured to Job?", min_value=0, max_value=int(count_l2), value=0, step=1, key=f"l2_grp_manufactured_{g}")
                            fields_l2 = st.multiselect("Fields to populate", options=multiselect_options_l2, default=["PartNo", "Description", "Quantity", "Cost"], key=f"l2_fields_{g}")
                            level2_config.append({"parent": parent_l1, "count": count_l2, "random": True, "fields": fields_l2, "manufactured_count": min(int(l2_grp_manufactured), count_l2)})
                        else:
                            level2_config.append({"parent": parent_l1, "count": count_l2, "random": False, "manual_rows": []})
                            for j in range(count_l2):
                                with st.expander(f"Level 2 part {j+1}"):
                                    cols = st.columns([2, 0.8, 2, 2, 1])  # narrow Qty/UM, wide Source
                                    with cols[0]:
                                        pno = st.text_input("PartNo *", key=f"l2_{g}_partno_{j}")
                                        level2_manual_partnos.append(pno)
                                        desc = st.text_input("Description", key=f"l2_{g}_desc_{j}")
                                    with cols[1]:
                                        qty = st.number_input("Quantity *", min_value=0.0, value=1.0, key=f"l2_{g}_qty_{j}", step=1.0, format="%.2f")
                                        um = st.text_input("UM", value="EA", key=f"l2_{g}_um_{j}")
                                    with cols[2]:
                                        src_label = st.selectbox("Source", options=SOURCE_LABELS, key=f"l2_{g}_src_{j}")
                                        cat_label = st.selectbox("Category", options=CATEGORY_LABELS, key=f"l2_{g}_cat_{j}")
                                    with cols[3]:
                                        pl = st.text_input("Productline", key=f"l2_{g}_pl_{j}", placeholder="")
                                        cost_l2 = st.number_input("Cost", min_value=0.0, value=None, step=0.01, format="%.2f", key=f"l2_{g}_cost_{j}")
                                        rev_l2_default = (parent_revision or "") if apply_revision_to_all else ""
                                        rev_l2 = st.text_input("Revision", key=f"l2_{g}_rev_{j}_ar{apply_revision_to_all}_p{parent_revision or ''}", value=rev_l2_default)
                                    with cols[4]:
                                        loc_l2_default = (parent_location or "") if apply_location_to_all else ""
                                        loc_l2 = st.text_input("Location", key=f"l2_{g}_loc_{j}_al{apply_location_to_all}_p{parent_location or ''}", value=loc_l2_default)
                                    with st.expander("Additional fields", expanded=False):
                                        e1, e2, e3 = st.columns(3)
                                        with e1:
                                            altdesc1_l2 = st.text_input("AltDescription1", key=f"l2_{g}_altdesc1_{j}", placeholder="")
                                            altdesc2_l2 = st.text_input("AltDescription2", key=f"l2_{g}_altdesc2_{j}", placeholder="")
                                            descextra_l2 = st.text_input("DescExtra", key=f"l2_{g}_descextra_{j}", placeholder="")
                                            issue_um_l2 = st.text_input("IssueUM", key=f"l2_{g}_issue_um_{j}", placeholder="")
                                            consumption_l2 = st.text_input("ConsumptionConv", key=f"l2_{g}_consumption_{j}", placeholder="")
                                            drawing_l2 = st.text_input("Drawing", key=f"l2_{g}_drawing_{j}", placeholder="")
                                            leadtime_l2 = st.text_input("Leadtime", key=f"l2_{g}_leadtime_{j}", placeholder="")
                                        with e2:
                                            memo1_l2 = st.text_input("Memo1", key=f"l2_{g}_memo1_{j}", placeholder="")
                                            memo2_l2 = st.text_input("Memo2", key=f"l2_{g}_memo2_{j}", placeholder="")
                                            sortcode_l2 = st.text_input("SortCode", key=f"l2_{g}_sortcode_{j}", placeholder="")
                                            tag_l2 = st.text_input("Tag", key=f"l2_{g}_tag_{j}", placeholder="")
                                        with e3:
                                            bomcomplete_l2 = st.text_input("BomComplete", key=f"l2_{g}_bomcomplete_{j}", placeholder="")
                                            bomcomments_l2 = st.text_input("BomComments", key=f"l2_{g}_bomcomments_{j}", placeholder="")
                                            router_l2 = st.text_input("Router", key=f"l2_{g}_router_{j}", placeholder="")
                                    level2_config[-1]["manual_rows"].append({
                                        "PartNo": pno, "Description": desc, "Quantity": qty, "UM": um,
                                        "Source": SOURCE_DISPLAY_TO_VALUE.get(src_label) or None,
                                        "Category": CATEGORY_DISPLAY_TO_VALUE.get(cat_label) or None,
                                        "Productline": pl, "Revision": rev_l2, "Location": loc_l2,
                                        "AltDescription1": altdesc1_l2 or None, "AltDescription2": altdesc2_l2 or None, "DescExtra": descextra_l2 or None,
                                        "IssueUM": issue_um_l2 or None, "ConsumptionConv": consumption_l2 or None,
                                        "Cost": round(cost_l2, 2) if cost_l2 is not None else None, "Drawing": drawing_l2 or None, "Leadtime": leadtime_l2 or None,
                                        "Memo1": memo1_l2 or None, "Memo2": memo2_l2 or None, "SortCode": sortcode_l2 or None, "Tag": tag_l2 or None,
                                        "BomComplete": bomcomplete_l2 or None, "BomComments": bomcomments_l2 or None, "Router": router_l2 or None,
                                        "Parent": parent_l1, "Level": 2, "Sequence": j + 1,
                                    })
            st.markdown("</div>", unsafe_allow_html=True)

    level2_partnos_for_l3 = [p for p in level2_manual_partnos if p]

    # Map Level 2 part numbers to Source (for blocking Level 3 when only one L2 parent and not manufactured)
    level2_partno_to_source = {}
    for cfg in level2_config:
        for r in cfg.get("manual_rows", []):
            pno = r.get("PartNo")
            if pno:
                level2_partno_to_source[pno] = r.get("Source")

    # Block Level 3 if no Level 2 part has a Manufactured Source (M or F)
    any_l2_manufactured = any(level2_partno_to_source.get(pno) in ("M", "F") for pno in level2_partnos_for_l3)
    block_l3 = bool(level2_partnos_for_l3) and not any_l2_manufactured
    # Level 3 parent options: only Level 2 parts that are Manufactured (M or F)
    level2_partnos_manufactured = [pno for pno in level2_partnos_for_l3 if level2_partno_to_source.get(pno) in ("M", "F")]

    # --- Level 3 Sub-Components: tree-style expander ---
    level3_config = []
    if level1_partnos:
        with st.expander("📁 Level 3 — Sub-components", expanded=True):
            st.markdown('<div style="margin-left: 3rem; border-left: 2px solid #888; padding-left: 0.75rem;">', unsafe_allow_html=True)
            st.caption("Add parts under a Level 2 part. You can randomly generate Level 3 for all Level 2 parts (including randomly generated ones), and/or add groups with a specific Level 2 parent.")
            st.info("**Random generation is capped at Level 3.** Deeper levels are not supported when using random generation.")
            if block_l3:
                st.warning(
                    "No Level 2 part has a **Manufactured** Source. At least one part must be **Manufactured to Stock** or **Manufactured to Job** to use as a parent for Level 3 Sub-Components. "
                    "Change a part's Source in the Level 2 section above to a Manufactured option."
                )
                num_l3_groups = st.number_input("Number of Level 3 groups", min_value=0, max_value=0, value=0, step=1, key="num_l3_groups", disabled=True)
            else:
                # Randomly generate sub-components for each Level 2 part
                with st.expander("Randomly generate Level 3 Sub-Components for each Manufactured Level 2 Part", expanded=False):
                    use_random_l3_per_l2 = st.checkbox("For each Manufactured Level 2 Part, randomly generate Level 3 Sub-Components", key="random_l3_per_l2")
                    if use_random_l3_per_l2:
                        count_random_l3 = st.number_input("Number of Level 3 parts per Level 2 part", min_value=1, max_value=20, value=2, key="count_random_l3")
                        l3_per_l2_manufactured = st.number_input("How many sub-components to have Source of Manufactured to Job?", min_value=0, max_value=int(count_random_l3), value=0, step=1, key="l3_per_l2_manufactured")
                        multiselect_options_l3 = [f for f in TEMPLATE_HEADERS if f not in ("Parent", "Sequence", "Level")]
                        fields_random_l3 = st.multiselect("Fields to populate (L3)", options=multiselect_options_l3, default=["PartNo", "Description", "Quantity", "Cost"], key="fields_random_l3")
                        random_l3_per_l2 = {"count": count_random_l3, "fields": fields_random_l3, "manufactured_count": min(int(l3_per_l2_manufactured), count_random_l3)}
                if level2_partnos_for_l3:
                    st.caption("Or add groups below (parent dropdown lists PartNos from your Level 2 manual entries).")
                elif not level2_partnos_for_l3:
                    st.info("You have no **manual** Level 2 parts. Use **\"Randomly generate Level 3 Sub-Components\"** above — when you generate the BOM, Level 3 parts will be created under every manufactured Level 2 part. To add specific Level 3 groups with a chosen parent, add at least one Level 2 part manually in section 3.")
                num_l3_groups = st.number_input("Number of Level 3 groups", min_value=0, max_value=20, value=0, step=1, key="num_l3_groups")
            random_all_l3 = False
            if not block_l3 and num_l3_groups > 0 and level2_partnos_manufactured:
                random_all_l3 = st.checkbox("Randomly generate all Level 3 groups. Note: Each group will have 2 Parts.", key="random_all_l3")
            multiselect_options_l3 = [f for f in TEMPLATE_HEADERS if f not in ("Parent", "Sequence", "Level")]
            if random_all_l3:
                l3_all_manufactured = st.number_input("How many sub-components per group to have Source of Manufactured to Job?", min_value=0, max_value=2, value=0, step=1, key="l3_all_manufactured")
                fields_all_l3 = st.multiselect("Fields to populate (all Level 3 groups)", options=multiselect_options_l3, default=["PartNo", "Description", "Quantity", "Cost"], key="fields_all_l3")
                for g in range(int(num_l3_groups)):
                    parent_l2 = level2_partnos_manufactured[g % len(level2_partnos_manufactured)]
                    level3_config.append({"parent": parent_l2, "count": 2, "random": True, "fields": fields_all_l3, "manufactured_count": min(int(l3_all_manufactured), 2)})
            else:
                for g in range(int(num_l3_groups)):
                    if not level2_partnos_for_l3:
                        st.info("**Level 3 groups** need at least one manual Level 2 part to choose as parent. Set **Number of Level 3 groups** to 0 and use **\"Randomly generate Level 3 Sub-Components\"** above — then click **Generate BOM** and Level 3 parts will be created under all manufactured Level 2 parts.")
                        break
                    with st.expander(f"Level 3 group {g+1}"):
                        parent_l2 = st.selectbox("Parent (Level 2 part)", level2_partnos_manufactured, index=0, key=f"l3_parent_{g}")
                        count_l3 = st.number_input("Number of parts", min_value=1, max_value=50, value=2, key=f"l3_count_{g}")
                        use_random_l3 = st.checkbox("Randomly generate this group", key=f"l3_random_{g}")
                        if use_random_l3:
                            l3_grp_manufactured = st.number_input("How many sub-components to have Source of Manufactured to Job?", min_value=0, max_value=int(count_l3), value=0, step=1, key=f"l3_grp_manufactured_{g}")
                            fields_l3 = st.multiselect("Fields to populate", options=multiselect_options_l3, default=["PartNo", "Description", "Quantity", "Cost"], key=f"l3_fields_{g}")
                            level3_config.append({"parent": parent_l2, "count": count_l3, "random": True, "fields": fields_l3, "manufactured_count": min(int(l3_grp_manufactured), count_l3)})
                        else:
                            level3_config.append({"parent": parent_l2, "count": count_l3, "random": False, "manual_rows": []})
                            for j in range(count_l3):
                                with st.expander(f"Level 3 part {j+1}"):
                                    cols = st.columns([2, 0.8, 2, 2, 1])  # narrow Qty/UM, wide Source
                                    with cols[0]:
                                        pno = st.text_input("PartNo *", key=f"l3_{g}_partno_{j}")
                                        desc = st.text_input("Description", key=f"l3_{g}_desc_{j}")
                                    with cols[1]:
                                        qty = st.number_input("Quantity *", min_value=0.0, value=1.0, key=f"l3_{g}_qty_{j}", step=1.0, format="%.2f")
                                        um = st.text_input("UM", value="EA", key=f"l3_{g}_um_{j}")
                                    with cols[2]:
                                        src_label = st.selectbox("Source", options=SOURCE_LABELS, key=f"l3_{g}_src_{j}")
                                        cat_label = st.selectbox("Category", options=CATEGORY_LABELS, key=f"l3_{g}_cat_{j}")
                                    with cols[3]:
                                        pl = st.text_input("Productline", key=f"l3_{g}_pl_{j}", placeholder="")
                                        cost_l3 = st.number_input("Cost", min_value=0.0, value=None, step=0.01, format="%.2f", key=f"l3_{g}_cost_{j}")
                                        rev_l3_default = (parent_revision or "") if apply_revision_to_all else ""
                                        rev_l3 = st.text_input("Revision", key=f"l3_{g}_rev_{j}_ar{apply_revision_to_all}_p{parent_revision or ''}", value=rev_l3_default)
                                    with cols[4]:
                                        loc_l3_default = (parent_location or "") if apply_location_to_all else ""
                                        loc_l3 = st.text_input("Location", key=f"l3_{g}_loc_{j}_al{apply_location_to_all}_p{parent_location or ''}", value=loc_l3_default)
                                    with st.expander("Additional fields", expanded=False):
                                        e1, e2, e3 = st.columns(3)
                                        with e1:
                                            altdesc1_l3 = st.text_input("AltDescription1", key=f"l3_{g}_altdesc1_{j}", placeholder="")
                                            altdesc2_l3 = st.text_input("AltDescription2", key=f"l3_{g}_altdesc2_{j}", placeholder="")
                                            descextra_l3 = st.text_input("DescExtra", key=f"l3_{g}_descextra_{j}", placeholder="")
                                            issue_um_l3 = st.text_input("IssueUM", key=f"l3_{g}_issue_um_{j}", placeholder="")
                                            consumption_l3 = st.text_input("ConsumptionConv", key=f"l3_{g}_consumption_{j}", placeholder="")
                                            drawing_l3 = st.text_input("Drawing", key=f"l3_{g}_drawing_{j}", placeholder="")
                                            leadtime_l3 = st.text_input("Leadtime", key=f"l3_{g}_leadtime_{j}", placeholder="")
                                        with e2:
                                            memo1_l3 = st.text_input("Memo1", key=f"l3_{g}_memo1_{j}", placeholder="")
                                            memo2_l3 = st.text_input("Memo2", key=f"l3_{g}_memo2_{j}", placeholder="")
                                            sortcode_l3 = st.text_input("SortCode", key=f"l3_{g}_sortcode_{j}", placeholder="")
                                            tag_l3 = st.text_input("Tag", key=f"l3_{g}_tag_{j}", placeholder="")
                                        with e3:
                                            bomcomplete_l3 = st.text_input("BomComplete", key=f"l3_{g}_bomcomplete_{j}", placeholder="")
                                            bomcomments_l3 = st.text_input("BomComments", key=f"l3_{g}_bomcomments_{j}", placeholder="")
                                            router_l3 = st.text_input("Router", key=f"l3_{g}_router_{j}", placeholder="")
                                    level3_config[-1]["manual_rows"].append({
                                        "PartNo": pno, "Description": desc, "Quantity": qty, "UM": um,
                                        "Source": SOURCE_DISPLAY_TO_VALUE.get(src_label) or None,
                                        "Category": CATEGORY_DISPLAY_TO_VALUE.get(cat_label) or None,
                                        "Productline": pl, "Revision": rev_l3, "Location": loc_l3,
                                        "AltDescription1": altdesc1_l3 or None, "AltDescription2": altdesc2_l3 or None, "DescExtra": descextra_l3 or None,
                                        "IssueUM": issue_um_l3 or None, "ConsumptionConv": consumption_l3 or None,
                                        "Cost": round(cost_l3, 2) if cost_l3 is not None else None, "Drawing": drawing_l3 or None, "Leadtime": leadtime_l3 or None,
                                        "Memo1": memo1_l3 or None, "Memo2": memo2_l3 or None, "SortCode": sortcode_l3 or None, "Tag": tag_l3 or None,
                                        "BomComplete": bomcomplete_l3 or None, "BomComments": bomcomments_l3 or None, "Router": router_l3 or None,
                                        "Parent": parent_l2, "Level": 3, "Sequence": j + 1,
                                    })
            st.markdown("</div>", unsafe_allow_html=True)

    st.divider()
    if st.button("Generate BOM", type="primary", use_container_width=True):
        errors = []
        if not parent_row:
            errors.append("Parent part (PartNo and Description) is required.")
        else:
            err = _validate_category_source(parent_row.get("Category"), parent_row.get("Source"), "Parent part")
            if err:
                errors.append(err)
        for idx, child in enumerate(child_data, 1):
            missing = [f for f in REQUIRED_FIELDS if child.get(f) in (None, "", [])]
            if missing:
                errors.append(f"Level 1 part #{idx}: Missing {', '.join(missing)}")
            else:
                err = _validate_category_source(child.get("Category"), child.get("Source"), f"Level 1 part #{idx} ({child.get('PartNo', '')})")
                if err:
                    errors.append(err)
        level2_rows = []
        level3_rows = []
        if not errors:
            # 1) Random Level 2 for each Manufactured Level 1 part (if option set)
            if random_l2_per_l1 and random_child_rows:
                fset = set(random_l2_per_l1["fields"]) | {"Parent", "Sequence", "Level", "PartNo", "Quantity"}
                count_r2 = random_l2_per_l1["count"]
                mfg_count_r2 = random_l2_per_l1.get("manufactured_count", 0)
                start_l2 = len(level2_rows)
                manufactured_l1 = [c for c in child_data if c.get("Source") in ("M", "F")]
                for child in manufactured_l1:
                    level2_rows.extend(random_child_rows(child["PartNo"], count_r2, fields_to_populate=fset, level=2))
                for i in range(len(manufactured_l1)):
                    batch = level2_rows[start_l2 + i * count_r2 : start_l2 + (i + 1) * count_r2]
                    _apply_random_row_defaults(batch, mfg_count_r2, apply_revision_to_all, apply_location_to_all, parent_revision, parent_location)
            # 2) Level 2 groups (manual or random)
            for cfg in level2_config:
                if cfg["random"] and random_child_rows:
                    fset = set(cfg["fields"]) | {"Parent", "Sequence", "Level", "PartNo", "Quantity"}
                    start_l2 = len(level2_rows)
                    level2_rows.extend(random_child_rows(cfg["parent"], cfg["count"], fields_to_populate=fset, level=2))
                    _apply_random_row_defaults(level2_rows[start_l2:], cfg.get("manufactured_count", 0), apply_revision_to_all, apply_location_to_all, parent_revision, parent_location)
                else:
                    for r in cfg.get("manual_rows", []):
                        row = {h: None for h in TEMPLATE_HEADERS}
                        for key in ("PartNo", "Description", "Quantity", "UM", "Source", "Category", "Productline", "Revision", "Location",
                                    "AltDescription1", "AltDescription2", "DescExtra", "IssueUM", "ConsumptionConv", "Cost", "Drawing", "Leadtime",
                                    "Memo1", "Memo2", "SortCode", "Tag", "BomComplete", "BomComments", "Router"):
                            if key in r and r[key] is not None:
                                row[key] = r[key]
                        row["UM"] = row.get("UM") or "EA"
                        row["IssueUM"] = r.get("IssueUM")
                        row["ConsumptionConv"] = r.get("ConsumptionConv")
                        row["Parent"] = cfg["parent"]
                        row["Level"] = 2
                        row["Sequence"] = r.get("Sequence", 0)
                        level2_rows.append(row)
            # 3) Random Level 3 for each Manufactured Level 2 part (if option set)
            if random_l3_per_l2 and random_child_rows and level2_rows:
                fset = set(random_l3_per_l2["fields"]) | {"Parent", "Sequence", "Level", "PartNo", "Quantity"}
                count_r3 = random_l3_per_l2["count"]
                mfg_count_r3 = random_l3_per_l2.get("manufactured_count", 0)
                start_l3 = len(level3_rows)
                manufactured_l2 = [r for r in level2_rows if r.get("Source") in ("M", "F")]
                for l2_row in manufactured_l2:
                    level3_rows.extend(random_child_rows(l2_row["PartNo"], count_r3, fields_to_populate=fset, level=3))
                for i in range(len(manufactured_l2)):
                    batch = level3_rows[start_l3 + i * count_r3 : start_l3 + (i + 1) * count_r3]
                    _apply_random_row_defaults(batch, mfg_count_r3, apply_revision_to_all, apply_location_to_all, parent_revision, parent_location)
            # 4) Level 3 groups (manual or random)
            for cfg in level3_config:
                if cfg["random"] and random_child_rows:
                    fset = set(cfg["fields"]) | {"Parent", "Sequence", "Level", "PartNo", "Quantity"}
                    start_l3 = len(level3_rows)
                    level3_rows.extend(random_child_rows(cfg["parent"], cfg["count"], fields_to_populate=fset, level=3))
                    _apply_random_row_defaults(level3_rows[start_l3:], cfg.get("manufactured_count", 0), apply_revision_to_all, apply_location_to_all, parent_revision, parent_location)
                else:
                    for r in cfg.get("manual_rows", []):
                        row = {h: None for h in TEMPLATE_HEADERS}
                        for key in ("PartNo", "Description", "Quantity", "UM", "Source", "Category", "Productline", "Revision", "Location",
                                    "AltDescription1", "AltDescription2", "DescExtra", "IssueUM", "ConsumptionConv", "Cost", "Drawing", "Leadtime",
                                    "Memo1", "Memo2", "SortCode", "Tag", "BomComplete", "BomComments", "Router"):
                            if key in r and r[key] is not None:
                                row[key] = r[key]
                        row["UM"] = row.get("UM") or "EA"
                        row["IssueUM"] = r.get("IssueUM")
                        row["ConsumptionConv"] = r.get("ConsumptionConv")
                        row["Parent"] = cfg["parent"]
                        row["Level"] = 3
                        row["Sequence"] = r.get("Sequence", 0)
                        level3_rows.append(row)
            # Apply sequence increment: Parent=0; Level 1 = 1*inc, 2*inc, ...; Level 2/3 reset and same pattern
            for idx, c in enumerate(child_data):
                c["Sequence"] = (idx + 1) * sequence_increment
            for idx, r in enumerate(level2_rows):
                r["Sequence"] = (idx + 1) * sequence_increment
            for idx, r in enumerate(level3_rows):
                r["Sequence"] = (idx + 1) * sequence_increment
            for r in level2_rows + level3_rows:
                missing = [f for f in REQUIRED_FIELDS if r.get(f) in (None, "", [])]
                if missing:
                    errors.append(f"Part {r.get('PartNo')}: Missing {', '.join(missing)}")
                else:
                    err = _validate_category_source(r.get("Category"), r.get("Source"), f"Part {r.get('PartNo', '')}")
                    if err:
                        errors.append(err)
            # Apply parent Revision and/or Location to all components when options are checked
            if apply_revision_to_all:
                rev_val = parent_revision or None
                for r in child_data:
                    r["Revision"] = rev_val
                for r in level2_rows:
                    r["Revision"] = rev_val
                for r in level3_rows:
                    r["Revision"] = rev_val
            if apply_location_to_all:
                loc_val = parent_location or None
                for r in child_data:
                    r["Location"] = loc_val
                for r in level2_rows:
                    r["Location"] = loc_val
                for r in level3_rows:
                    r["Location"] = loc_val
        if errors:
            for e in errors:
                st.error(e)
        else:
            output_wb = openpyxl.load_workbook(template_path)
            output_ws = output_wb["Template"]
            output_headers = load_template_headers(output_ws)
            if output_ws.max_row > 1:
                output_ws.delete_rows(2, output_ws.max_row)
            # Parent row: set Parent = PartNo for Level 0 so BOM Compare accepts it ("part must have a parent part number")
            parent_values = []
            for h in output_headers:
                if h == "Parent":
                    parent_values.append(parent_partno)  # top-level part uses its own PartNo as Parent
                else:
                    v = parent_row.get(h, None)
                    parent_values.append("" if v is None else v)
            output_ws.append(parent_values)
            # Avoid self-reference: if any row has PartNo == Parent, clear Parent so BOM Compare won't error
            def clear_self_parent(row):
                if row.get("PartNo") and row.get("Parent") and str(row["PartNo"]).strip() == str(row["Parent"]).strip():
                    row = dict(row)
                    row["Parent"] = ""
                return row
            for child in child_data:
                append_row_by_headers(output_ws, output_headers, clear_self_parent(child))
            for row in level2_rows:
                append_row_by_headers(output_ws, output_headers, clear_self_parent(row))
            for row in level3_rows:
                append_row_by_headers(output_ws, output_headers, clear_self_parent(row))
            output_buffer = io.BytesIO()
            output_wb.save(output_buffer)
            output_buffer.seek(0)
            total_parts = 1 + len(child_data) + len(level2_rows) + len(level3_rows)
            st.success("BOM generated successfully. Parent (Level 0), then Level 1, Level 2, and Level 3.")
            filename = f"BOM_{parent_partno}_{total_parts}_parts.xlsx"
            st.download_button(
                label="Download BOM",
                data=output_buffer,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            with st.expander("BOM Summary"):
                st.write("**Parent (Level 0):**", parent_partno)
                st.write("**Level 1:**", len(child_data))
                st.write("**Level 2:**", len(level2_rows))
                st.write("**Level 3:**", len(level3_rows))
                st.write("**Total rows:**", total_parts)


main()
